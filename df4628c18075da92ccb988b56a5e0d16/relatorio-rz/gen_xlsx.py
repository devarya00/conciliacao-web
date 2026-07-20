# Gera relatorio-gerencial.xlsx a partir de relatorio.db.
# Abas de dados (Dados_Balancete, Dados_Acumuladores) carregam a base extraída;
# todas as demais abas calculam seus números por fórmula (INDEX/MATCH, SUMIFS, SUM)
# — inclusive o saldo atual do balancete (= anterior + débito − crédito).
import sqlite3
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = sys.argv[1] if len(sys.argv) > 1 else "relatorio.db"
OUT = sys.argv[2] if len(sys.argv) > 2 else "relatorio-gerencial.xlsx"

NAVY = "0D366B"
NAVY2 = "104281"
WASH = "F4F6FA"
GRID = "E1E0D9"
MUTED = "898781"
INK2 = "52514E"
BLUE = "2A78D6"

FMT_BRL = '#,##0.00;[Red](#,##0.00)'
FMT_PCT = '0.0%'

thin = Side(style="thin", color=GRID)
medium = Side(style="medium", color="C3C2B7")

f_title = Font(size=16, bold=True, color="FFFFFF")
f_sub = Font(size=10, color="CDE2FB")
f_h = Font(size=11, bold=True, color=NAVY)
f_colhdr = Font(size=8, bold=True, color="FFFFFF")
f_lbl = Font(size=8, bold=True, color=MUTED)
f_group = Font(bold=True)
f_total = Font(bold=True)
f_note = Font(size=8, color=INK2)
fill_hdr = PatternFill("solid", fgColor=NAVY)
fill_wash = PatternFill("solid", fgColor=WASH)


def col_headers(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = f_colhdr
        c.fill = fill_hdr
        c.alignment = Alignment(horizontal="left" if i <= 2 or h in ("Descrição", "Conta", "Linha", "Fornecedor", "Verificação", "Grupo", "Acumulador", "Seção", "Coluna", "Nat.") else "right", vertical="center")
        c.border = Border(bottom=medium)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 18


def money(ws, cell, formula_or_value, bold=False, fill=None):
    c = ws[cell]
    c.value = formula_or_value
    c.number_format = FMT_BRL
    if bold:
        c.font = f_total
    if fill:
        c.fill = fill
    return c


con = sqlite3.connect(DB)
meta = dict(
    con.execute(
        "SELECT 'x', empresa FROM relatorio WHERE tipo='balancete'"
    ).fetchall()
)
emp, cnpj, p0, p1, emis = con.execute(
    "SELECT empresa, cnpj, periodo_ini, periodo_fim, emissao FROM relatorio WHERE tipo='balancete'"
).fetchone()

bal = con.execute(
    "SELECT codigo, pai, nivel, descricao, saldo_atual_dc, saldo_anterior_signed,"
    " debito, credito FROM balancete ORDER BY rowid"
).fetchall()

acu = con.execute(
    """SELECT s.nome, a.codigo, a.descricao, v.coluna, v.valor
       FROM acumulador a
       JOIN acumulador_secao s ON s.id = a.secao_id
       JOIN acumulador_valor v ON v.acumulador_id = a.id
       WHERE a.is_total = 0 ORDER BY a.id, v.rowid"""
).fetchall()

acu_tot = con.execute(
    """SELECT s.nome, v.valor FROM acumulador a
       JOIN acumulador_secao s ON s.id=a.secao_id
       JOIN acumulador_valor v ON v.acumulador_id=a.id AND v.coluna='Vlr Contábil'
       WHERE a.is_total=1"""
).fetchall()

forn = con.execute(
    """SELECT b.codigo, b.descricao FROM balancete b
       WHERE b.pai=(SELECT codigo FROM balancete WHERE descricao='FORNECEDORES NACIONAIS')
       ORDER BY b.saldo_atual DESC"""
).fetchall()

wb = Workbook()

# ================================================================ Dados_Balancete
ws = wb.active
ws.title = "Dados_Balancete"
ws.sheet_properties.tabColor = MUTED
col_headers(
    ws, 1,
    ["Código", "Pai", "Nível", "Descrição", "Nat.", "Saldo anterior",
     "Débito", "Crédito", "Saldo atual"],
    [9, 7, 7, 46, 6, 15, 13, 13, 15],
)
for r, (cod, pai, niv, desc, dc, sa, deb, cred) in enumerate(bal, 2):
    ws.cell(row=r, column=1, value=cod)
    ws.cell(row=r, column=2, value=pai)
    ws.cell(row=r, column=3, value=niv)
    d = ws.cell(row=r, column=4, value=desc)
    d.alignment = Alignment(indent=niv - 1)
    ws.cell(row=r, column=5, value=dc)
    money(ws, f"F{r}", sa)
    money(ws, f"G{r}", deb)
    money(ws, f"H{r}", cred)
    money(ws, f"I{r}", f"=F{r}+G{r}-H{r}", bold=(niv == 1))
    if niv == 1:
        for cc in "ABCDEFGHI":
            ws[f"{cc}{r}"].fill = fill_wash
            ws[f"{cc}{r}"].font = f_group
nbal = len(bal) + 1  # última linha
ws.freeze_panes = "A2"
note = ws.cell(row=nbal + 2, column=1,
               value="Convenção de sinal: devedor positivo, credor negativo. "
                     "Saldo atual é fórmula: anterior + débito − crédito.")
note.font = f_note

# ================================================================ Dados_Acumuladores
ws = wb.create_sheet("Dados_Acumuladores")
ws.sheet_properties.tabColor = MUTED
col_headers(ws, 1, ["Seção", "Código", "Acumulador", "Coluna", "Valor"],
            [12, 9, 38, 15, 14])
for r, (sec, cod, desc, colname, val) in enumerate(acu, 2):
    ws.cell(row=r, column=1, value=sec)
    ws.cell(row=r, column=2, value=cod)
    ws.cell(row=r, column=3, value=desc)
    ws.cell(row=r, column=4, value=colname)
    money(ws, f"E{r}", val)
nacu = len(acu) + 1
ws.freeze_panes = "A2"

BAL = "Dados_Balancete"
ACU = "Dados_Acumuladores"


def sa(cod):   # saldo anterior (assinado)
    return f"INDEX({BAL}!$F$2:$F${nbal},MATCH({cod},{BAL}!$A$2:$A${nbal},0))"


def deb(cod):
    return f"INDEX({BAL}!$G$2:$G${nbal},MATCH({cod},{BAL}!$A$2:$A${nbal},0))"


def cred(cod):
    return f"INDEX({BAL}!$H$2:$H${nbal},MATCH({cod},{BAL}!$A$2:$A${nbal},0))"


def sf(cod):   # saldo atual (assinado, célula-fórmula da base)
    return f"INDEX({BAL}!$I$2:$I${nbal},MATCH({cod},{BAL}!$A$2:$A${nbal},0))"


# ================================================================ DRE
ws = wb.create_sheet("DRE")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "Demonstração do resultado"
ws["A1"].font = f_h
ws["A2"] = f"{emp} · movimentos de {p0} a {p1} e saldos acumulados do exercício"
ws["A2"].font = f_note
col_headers(ws, 4, ["Linha", "Bimestre (mov.)", "% rec. líq.", "Acumulado (saldo)"],
            [44, 17, 12, 17])

dre_rows = [
    # (rótulo, fórmula bimestre, fórmula acumulado, tipo)
    ("Receita de prestação de serviços", f"={cred(404)}-{deb(404)}", f"=-{sf(404)}", ""),
    ("(−) Deduções — Simples Nacional", f"=-({deb(413)}-{cred(413)})", f"=-{sf(413)}", "sub"),
    ("Receita líquida", "=B5+B6", "=D5+D6", "group"),
    ("(−) Custos — combustíveis e energia", f"=-({deb(500)}-{cred(500)})", f"=-{sf(500)}", ""),
    ("Resultado bruto", "=B7+B8", "=D7+D8", "group"),
    ("(−) Despesas com vendas", f"=-({deb(296)}-{cred(296)})", f"=-{sf(296)}", ""),
    ("(−) Despesas administrativas", f"=-({deb(329)}-{cred(329)})", f"=-{sf(329)}", ""),
    ("Resultado do período", "=B9+B10+B11", "=D9+D10+D11", "total"),
]
for i, (lbl, fb, fa, kind) in enumerate(dre_rows):
    r = 5 + i
    c = ws.cell(row=r, column=1, value=lbl)
    money(ws, f"B{r}", fb)
    p = ws.cell(row=r, column=3, value=f"=IF($B$7=0,0,B{r}/$B$7)")
    p.number_format = FMT_PCT
    money(ws, f"D{r}", fa)
    if kind in ("group", "total"):
        for cc in "ABCD":
            ws[f"{cc}{r}"].font = f_total
            if kind == "group":
                ws[f"{cc}{r}"].fill = fill_wash
    if kind == "total":
        for cc in "ABCD":
            ws[f"{cc}{r}"].border = Border(top=medium)
    if kind == "sub":
        c.font = Font(color=INK2)
        c.alignment = Alignment(indent=1)
DRE_RES_BIM = "DRE!$B$12"
DRE_RES_ACU = "DRE!$D$12"
DRE_REC_BIM = "DRE!$B$5"

# ================================================================ Balanço
ws = wb.create_sheet("Balanço")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "Balanço patrimonial"
ws["A1"].font = f_h
ws["A2"] = f"Posição em {p1} · valores em R$"
ws["A2"].font = f_note
col_headers(ws, 4, ["Conta", "Saldo anterior", "Saldo atual"], [40, 17, 17])

bal_rows = [
    ("ATIVO", None, None, "h"),
    ("Ativo circulante", f"={sa(2)}", f"={sf(2)}", "group"),
    ("Disponível (caixa)", f"={sa(3)}", f"={sf(3)}", "sub"),
    ("Outros créditos", f"={sa(18)}", f"={sf(18)}", "sub"),
    ("Estoque — mercadorias e insumos", f"={sa(53)}", f"={sf(53)}", "sub"),
    ("Ativo não circulante", f"={sa(501)}", f"={sf(501)}", "group"),
    ("Veículos", f"={sa(120)}", f"={sf(120)}", "sub"),
    ("(−) Depreciações acumuladas", f"={sa(125)}", f"={sf(125)}", "sub"),
    ("Total do ativo", "=B6+B10", "=C6+C10", "total"),
    ("", None, None, "blank"),
    ("PASSIVO E PATRIMÔNIO LÍQUIDO", None, None, "h"),
    ("Passivo circulante", f"=-{sa(150)}", f"=-{sf(150)}", "group"),
    ("Fornecedores", f"=-{sa(164)}", f"=-{sf(164)}", "sub"),
    ("Obrigações tributárias", f"=-{sa(169)}", f"=-{sf(169)}", "sub"),
    ("Obrigações trabalhistas e previdenciárias", f"=-{sa(185)}", f"=-{sf(185)}", "sub"),
    ("Patrimônio líquido efetivo", "=B21+B22+B23", "=C21+C22+C23", "group"),
    ("Capital social subscrito", f"=-{sa(243)}", f"=-{sf(243)}", "sub"),
    ("(−) Lucros/prejuízos acumulados", f"=-{sa(264)}", f"=-{sf(264)}", "sub"),
    ("(−) Resultado do exercício não encerrado",
     f"=-({sa(269)}+{sa(402)})", f"=-({sf(269)}+{sf(402)})", "sub"),
    ("Total passivo + PL", "=B16+B20", "=C16+C20", "total"),
]
for i, (lbl, fb, fc, kind) in enumerate(bal_rows):
    r = 5 + i
    c = ws.cell(row=r, column=1, value=lbl)
    if fb:
        money(ws, f"B{r}", fb)
        money(ws, f"C{r}", fc)
    if kind == "h":
        c.font = Font(bold=True, color=NAVY, size=10)
    elif kind in ("group", "total"):
        for cc in "ABC":
            ws[f"{cc}{r}"].font = f_total
            if kind == "group":
                ws[f"{cc}{r}"].fill = fill_wash
            else:
                ws[f"{cc}{r}"].border = Border(top=medium)
    elif kind == "sub":
        c.font = Font(color=INK2)
        c.alignment = Alignment(indent=1)
BAL_ATIVO = "Balanço!$C$13"
BAL_PL = "Balanço!$C$20"
BAL_PASSIVO_PL = "Balanço!$C$24"

# ================================================================ Fornecedores
ws = wb.create_sheet("Fornecedores")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "Fornecedores nacionais — contas a pagar"
ws["A1"].font = f_h
ws["A2"] = "Saldos por fórmula a partir de Dados_Balancete; ordenado pelo saldo em " + p1
ws["A2"].font = f_note
col_headers(ws, 4,
            ["Código", "Fornecedor", "Saldo anterior", "Compras (créd.)",
             "Pagamentos (déb.)", "Saldo atual", "% do total"],
            [9, 42, 15, 15, 15, 15, 10])
first, last = 5, 5 + len(forn) - 1
tot_row = last + 1
for i, (cod, nome) in enumerate(forn):
    r = first + i
    ws.cell(row=r, column=1, value=cod)
    ws.cell(row=r, column=2, value=nome)
    money(ws, f"C{r}", f"=-{sa(f'$A{r}')}")
    money(ws, f"D{r}", f"={cred(f'$A{r}')}")
    money(ws, f"E{r}", f"={deb(f'$A{r}')}")
    money(ws, f"F{r}", f"=C{r}+D{r}-E{r}")
    p = ws.cell(row=r, column=7, value=f"=IF($F${tot_row}=0,0,F{r}/$F${tot_row})")
    p.number_format = FMT_PCT
ws.cell(row=tot_row, column=2, value="Total (soma das linhas)").font = f_total
for cc in "CDEF":
    money(ws, f"{cc}{tot_row}", f"=SUM({cc}{first}:{cc}{last})", bold=True)
    ws[f"{cc}{tot_row}"].border = Border(top=medium)
chk = ws.cell(row=tot_row + 1, column=2,
              value="Conferência — saldo sintético (conta 165):")
chk.font = f_note
money(ws, f"F{tot_row + 1}", f"=-{sf(165)}")
ws[f"F{tot_row + 1}"].font = f_note
ws.freeze_panes = "A5"
FORN_TOTAL = f"Fornecedores!$F${tot_row}"

# ================================================================ Fiscal
ws = wb.create_sheet("Fiscal")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "Movimento fiscal por acumulador"
ws["A1"].font = f_h
ws["A2"] = "Valor contábil por SUMIFS sobre Dados_Acumuladores"
ws["A2"].font = f_note


def sumifs_acu(sec, cod):
    return (f"=SUMIFS({ACU}!$E$2:$E${nacu},{ACU}!$A$2:$A${nacu},\"{sec}\","
            f"{ACU}!$D$2:$D${nacu},\"Vlr Contábil\",{ACU}!$B$2:$B${nacu},{cod})")


def sumifs_sec(sec):
    return (f"=SUMIFS({ACU}!$E$2:$E${nacu},{ACU}!$A$2:$A${nacu},\"{sec}\","
            f"{ACU}!$D$2:$D${nacu},\"Vlr Contábil\")")


col_headers(ws, 4, ["Cód", "Acumulador", "Vlr contábil"], [8, 42, 16])
fiscal = [
    ("ENTRADAS", [(119, "Compra de combustível a prazo"),
                  (401, "Serviços tomados a prazo"),
                  (107, "Compra de uso e consumo a prazo")]),
    ("SERVIÇOS", [(904, "P. serviço à vista — ISS retido"),
                  (900, "P. serviço à vista — ISS no município")]),
]
r = 5
FISCAL_TOT = {}
for sec, rows in fiscal:
    c = ws.cell(row=r, column=1, value=sec)
    c.font = f_group
    for cc in "ABC":
        ws[f"{cc}{r}"].fill = fill_wash
    r += 1
    for cod, desc in rows:
        ws.cell(row=r, column=1, value=cod)
        ws.cell(row=r, column=2, value=desc)
        money(ws, f"C{r}", sumifs_acu(sec, cod))
        r += 1
    ws.cell(row=r, column=2, value=f"Total {sec.lower()}").font = f_total
    money(ws, f"C{r}", sumifs_sec(sec), bold=True)
    ws[f"C{r}"].border = Border(top=medium)
    FISCAL_TOT[sec] = f"Fiscal!$C${r}"
    r += 2

# ================================================================ Verificações
ws = wb.create_sheet("Verificações")
ws.sheet_properties.tabColor = "D03B3B"
ws["A1"] = "Verificações de integridade"
ws["A1"].font = f_h
ws["A2"] = "Toda linha deve exibir OK — diferenças calculadas por fórmula."
ws["A2"].font = f_note
col_headers(ws, 4, ["Verificação", "Valor A", "Valor B", "Diferença", "Status"],
            [52, 16, 16, 12, 8])

soma_filhos = (f"SUMIFS({BAL}!$I$2:$I${nbal},{BAL}!$B$2:$B${nbal},{{}})")
checks = [
    ("Ativo = Passivo + PL (saldo atual)", f"={BAL_ATIVO}", f"={BAL_PASSIVO_PL}"),
    ("Resultado do bimestre (DRE) = variação patrimonial (A−P exigível)",
     f"={DRE_RES_BIM}",
     f"=({sf(1)}+{sf(149)})-({sa(1)}+{sa(149)})"),
    ("Resultado acumulado (DRE) = contas de resultado do balancete",
     f"={DRE_RES_ACU}", f"=-({sf(269)}+{sf(402)})"),
    ("Fornecedores: soma das linhas = conta sintética 165",
     f"={FORN_TOTAL}", f"=-{sf(165)}"),
    ("Receita bruta (DRE) = total de serviços por acumulador",
     f"={DRE_REC_BIM}", f"={FISCAL_TOT['SERVIÇOS']}"),
    ("Ativo: soma dos filhos = conta 1",
     f"={soma_filhos.format(1)}", f"={sf(1)}"),
    ("Passivo: soma dos filhos = conta 149",
     f"={soma_filhos.format(149)}", f"={sf(149)}"),
    ("Total de entradas por acumulador (extraído: 76.674,50)",
     f"={FISCAL_TOT['ENTRADAS']}", str(dict(acu_tot).get("ENTRADAS", 0)).replace(",", ".")),
]
for i, (lbl, fa, fb) in enumerate(checks):
    r = 5 + i
    ws.cell(row=r, column=1, value=lbl)
    money(ws, f"B{r}", fa)
    money(ws, f"C{r}", float(fb) if not str(fb).startswith("=") else fb)
    money(ws, f"D{r}", f"=B{r}-C{r}")
    st = ws.cell(row=r, column=5, value=f'=IF(ABS(D{r})<0.01,"OK","ERRO")')
    st.font = Font(bold=True)
    st.alignment = Alignment(horizontal="center")

# ================================================================ Início
ws = wb.create_sheet("Início", 0)
ws.sheet_properties.tabColor = NAVY
ws.sheet_view.showGridLines = False
for rng in ("A1:F1", "A2:F2", "A3:F3"):
    ws.merge_cells(rng)
for rr in (1, 2, 3):
    for cc in range(1, 7):
        ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=NAVY)
ws["A1"] = "Relatório Gerencial Contábil"
ws["A1"].font = Font(size=9, bold=True, color="9EC5F4")
ws["A2"] = emp
ws["A2"].font = f_title
ws["A3"] = f"CNPJ {cnpj} · Período {p0} – {p1} · dados emitidos em {emis}"
ws["A3"].font = f_sub
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 26
ws.row_dimensions[3].height = 20
for i, w in enumerate([34, 17, 4, 34, 17, 6], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

kpis = [
    ("Receita bruta do período", f"={DRE_REC_BIM}", 5),
    ("Resultado do período", f"={DRE_RES_BIM}", 7),
    ("Resultado acumulado do exercício", f"={DRE_RES_ACU}", 9),
    ("Total do ativo", f"={BAL_ATIVO}", 11),
    ("Fornecedores a pagar", f"={FORN_TOTAL}", 13),
    ("Patrimônio líquido efetivo", f"={BAL_PL}", 15),
]
for lbl, fx, r in kpis:
    ws.cell(row=r, column=1, value=lbl).font = f_lbl
    c = money(ws, f"B{r}", fx, bold=True)
    c.font = Font(size=13, bold=True)
    ws[f"A{r}"].border = Border(bottom=thin)
    ws[f"B{r}"].border = Border(bottom=thin)

guide = [
    ("Como este arquivo funciona", None),
    ("Dados_Balancete e Dados_Acumuladores", "base extraída dos PDFs do Domínio (SQLite)."),
    ("DRE · Balanço · Fornecedores · Fiscal", "todos os números são fórmulas sobre a base."),
    ("Verificações", "conferências de integridade — tudo deve exibir OK."),
]
r = 5
for t, d in guide:
    ws.cell(row=r, column=4, value=t).font = f_h if d is None else Font(bold=True, size=9)
    if d:
        ws.cell(row=r + 1, column=4, value=d).font = f_note
        r += 2
    else:
        r += 1
ws.cell(row=17, column=1,
        value="Rezende Contabilidade Empresarial e Tributária · gerado por gen_xlsx.py a partir de relatorio.db").font = f_note

wb.save(OUT)
print(f"gravado: {OUT}")
